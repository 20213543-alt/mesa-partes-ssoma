import os
import smtplib
import requests
import traceback
from typing import List, Optional
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from fastapi import FastAPI, Form
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse
import openpyxl

app = FastAPI()

# Enlace de tu Google Apps Script
GOOGLE_WEBHOOK_URL = "https://script.google.com/macros/s/AKfycbxaliz82ArStXwK5OH2lAn_wK0rp23CIvWy4cglATNt5AhV90VeucsJ7GrB1sFHYANhRw/exec"
PLANTILLA_PATH = "INFORME DE ACCIDENTES (1).xlsx"

# Configuración de correo
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = os.getenv("SENDER_EMAIL", "")
SENDER_PASSWORD = os.getenv("SENDER_PASSWORD", "")

CORREOS_NOTIFICACION = ["chanelone14@gmail.com", "20213543@aloe.ulima.edu.pe"]

def enviar_correo_notificacion(asunto: str, cuerpo: str):
    """Envía la notificación por correo de fondo con timeout seguro"""
    if not SENDER_EMAIL or not SENDER_PASSWORD:
        print("Variables de correo no configuradas en Render. Omitiendo envío de email.")
        return

    try:
        msg = MIMEMultipart()
        msg['From'] = SENDER_EMAIL
        msg['To'] = ", ".join(CORREOS_NOTIFICACION)
        msg['Subject'] = asunto
        msg.attach(MIMEText(cuerpo, 'plain'))

        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT, timeout=5)
        server.starttls()
        server.login(SENDER_EMAIL, SENDER_PASSWORD)
        server.sendmail(SENDER_EMAIL, CORREOS_NOTIFICACION, msg.as_string())
        server.quit()
        print("Notificación de correo enviada exitosamente.")
    except Exception as e:
        print(f"No se pudo enviar el correo (omitido para no pausar el flujo): {e}")

@app.get("/", response_class=HTMLResponse)
def home():
    if os.path.exists("index.html"):
        with open("index.html", "r", encoding="utf-8") as f:
            return f.read()
    return "<h1>Servidor SSOMA activo</h1>"

@app.post("/enviar-reporte")
def enviar_reporte(
    tipo_informe: str = Form("FINAL"),
    
    # Campos Preliminar
    tipo_evento_pre: Optional[str] = Form(None),
    fecha_evento_pre: Optional[str] = Form(None),
    hora_ocurrencia_pre: Optional[str] = Form(None),
    lugar_ocurrencia_pre: Optional[str] = Form(None),
    fecha_reporte_pre: Optional[str] = Form(None),
    trabajo_realizaba_pre: Optional[str] = Form(None),
    nombre_lesionado_pre: Optional[str] = Form(None),
    edad_lesionado_pre: Optional[str] = Form(None),
    cargo_lesionado_pre: Optional[str] = Form(None),
    breve_descripcion_pre: Optional[str] = Form(None),
    resp_nombre_pre: Optional[str] = Form(None),

    # Campos Final
    fin_fecha_evento: Optional[str] = Form(None),
    fin_hora_evento: Optional[str] = Form(None),
    fin_lugar_exacto: Optional[str] = Form(None),
    fin_tipo_evento: Optional[str] = Form(None),
    que_sucedio: Optional[str] = Form(None),
    trab_paterno: List[str] = Form([]),
    trab_materno: List[str] = Form([]),
    trab_nombres: List[str] = Form([]),
    trab_dni: List[str] = Form([])
):
    try:
        # Cargar plantilla base en Excel para la descarga del usuario
        if os.path.exists(PLANTILLA_PATH):
            wb = openpyxl.load_workbook(PLANTILLA_PATH)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active

        if tipo_informe == "PRELIMINAR":
            ws['C4'] = tipo_evento_pre or ""
            ws['C6'] = fecha_evento_pre or ""
            ws['J6'] = hora_ocurrencia_pre or ""
            ws['C7'] = lugar_ocurrencia_pre or ""
            ws['J7'] = fecha_reporte_pre or ""
            ws['C8'] = trabajo_realizaba_pre or ""
            ws['C10'] = nombre_lesionado_pre or ""
            ws['E10'] = edad_lesionado_pre or ""
            ws['F10'] = cargo_lesionado_pre or ""
            ws['C12'] = breve_descripcion_pre or ""
            ws['C20'] = resp_nombre_pre or ""
            
            asunto = f"NUEVO REGISTRO SSOMA: Informe Preliminar - {nombre_lesionado_pre or 'Anónimo'}"
            cuerpo = f"Se ha registrado un Informe Preliminar.\nTrabajador: {nombre_lesionado_pre}\nFecha: {fecha_evento_pre}"
            nombre_archivo = "Informe_Preliminar.xlsx"

        else:
            paterno = trab_paterno[0] if trab_paterno else ""
            materno = trab_materno[0] if trab_materno else ""
            nombres = trab_nombres[0] if trab_nombres else ""
            dni = trab_dni[0] if trab_dni else ""
            nombre_completo = f"{paterno} {materno} {nombres}".strip()

            # Llenar Excel individual de descarga
            ws['C30'] = fin_fecha_evento or ""
            ws['C31'] = fin_hora_evento or ""
            ws['R30'] = fin_lugar_exacto or ""

            # 1. ENVIAR A GOOGLE SHEETS (Las 11 columnas del Reporte Final)
            datos_envio = {
                "fin_fecha_evento": fin_fecha_evento,
                "fin_hora_evento": fin_hora_evento,
                "fin_lugar_exacto": fin_lugar_exacto,
                "fin_tipo_evento": fin_tipo_evento,
                "trab_nombres": nombres,
                "trab_paterno": paterno,
                "trab_materno": materno,
                "trab_dni": dni,
                "que_sucedio": que_sucedio
            }

            try:
                requests.post(GOOGLE_WEBHOOK_URL, json=datos_envio, timeout=8)
            except Exception as err_sheet:
                print(f"Error al enviar a Google Sheets: {err_sheet}")

            asunto = f"NUEVO REGISTRO SSOMA: Informe Final - {nombre_completo or 'Anónimo'}"
            cuerpo = f"Se ha registrado un Informe Final de Accidente.\nTrabajador: {nombre_completo}\nDNI: {dni}\nFecha: {fin_fecha_evento}"
            nombre_archivo = "Informe_Final.xlsx"

        # Guardar archivo individual para descarga inmediata
        wb.save(nombre_archivo)
        
        # 2. ENVIAR CORREO DE NOTIFICACIÓN A AMBOS DESTINATARIOS
        enviar_correo_notificacion(asunto, cuerpo)

        return FileResponse(
            path=nombre_archivo,
            filename=nombre_archivo,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        error_detallado = traceback.format_exc()
        print(error_detallado)
        return JSONResponse(status_code=500, content={"error": str(e), "trace": error_detallado})
